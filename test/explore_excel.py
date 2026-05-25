import openpyxl
from openpyxl.utils import get_column_letter

filepath = r"D:\claude-code-coding-test	est月分析底稿.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=True)

print("=" * 100)
print("1. ALL SHEET NAMES ({} total):".format(len(wb.sheetnames)))
print("=" * 100)
for i, name in enumerate(wb.sheetnames, 1):
    print("  {:2d}. [{}]".format(i, name))

wb.close()